#!/usr/bin/env python3
"""Merge SMART curve data from Waterbot markdown + RPS sizing XLSX into families.json."""
import json
import re
from pathlib import Path

import openpyxl

PROJ = Path(__file__).resolve().parents[1]
FAM_JSON = PROJ / "data" / "families.json"
MD = Path(
    r"C:\Users\Matthew\Proton Drive\matticus.headquarters\My files\GPT - Codex"
    r"\Waterbot 2.0\Reference Sheets\RPS Pump Curve SMART Export - Phase 1-3.md"
)
XLSX = Path(
    r"C:\Users\Matthew\Proton Drive\matticus.headquarters\My files\GPT - Codex"
    r"\Waterbot 2.0\Reference Sheets\RPS Water Pumps Schedule - Sizing Sheet.xlsx"
)

# XLSX wins on these (flagged markdown rows or missing from web scrape)
XLSX_WINS = {
    "90RPS100",
    "230RPS100",
    "300RPS150",
    "400RPS200",
    "400RPS250",
    "90RPS200",
    "90RPS250",
    "300RPS100",
}

HEAD_LABELS = {
    "05": "1/2 HP",
    "07": "3/4 HP",
    "10": "1 HP",
    "15": "1.5 HP",
    "20": "2 HP",
    "30": "3 HP",
    "50": "5 HP",
}
HEAD_COLORS = {
    "05": "#2f7fd0",
    "07": "#cf3b34",
    "10": "#e0a800",
    "15": "#2c9b3f",
    "20": "#e07b00",
    "30": "#008080",
    "50": "#6eb5d0",
}

ST_SHEETS = {
    "5RPS Sizing+S": "05",
    "7RPS Sizing": "07",
    "10RPS Sizing+S": "10",
    "13RPS Sizing": "13",
    "18RPS Sizing+S": "18",
    "25RPS Sizing+S": "25",
}

ST_FAMILY_HEADS = {
    "05RPS": ["05", "07", "10", "15", "20"],
    "07RPS": ["05", "07", "10", "15", "20", "30"],
    "10RPS": ["05", "07", "10", "15", "20", "30", "50"],
    "25RPS": ["05", "07", "10", "15", "20", "30", "50"],
}

ST_13_18 = (
    [("13", h) for h in ("05", "07", "10", "15", "20", "30", "50")]
    + [("18", h) for h in ("05", "07", "10", "15", "20", "30", "50")]
)


def norm_id(raw):
    s = str(raw).strip().replace("-S", "").replace(" ", "")
    m = re.match(r"^(\d+)RPS(.+)$", s)
    if m and len(m.group(1)) == 1:
        s = f"0{m.group(1)}RPS{m.group(2)}"
    return s


def parse_markdown(path):
    text = path.read_text(encoding="utf-8")
    out = {}
    for m in re.finditer(r"### (\S+) \(", text):
        model_id = m.group(1)
        chunk = text[m.end() : m.end() + 900]
        head = chunk.split("###")[0]
        if "NO PUBLISHED CURVE" in head.split("```")[0]:
            continue
        dm = re.search(r'"data": (\[\[.*?\]\])', head, re.DOTALL)
        if dm:
            out[model_id] = json.loads(dm.group(1))
    return out


def add_point(store, mid, tdh, gpm):
    if gpm is None or gpm <= 0:
        return
    store.setdefault(mid, []).append([tdh, gpm if gpm != int(gpm) else int(gpm)])


def dedupe_sort(rows):
    seen = {}
    for tdh, gpm in rows:
        seen[float(tdh)] = gpm
    return [[t, seen[t]] for t in sorted(seen)]


def parse_big_sheet(rows):
    hdr_i = None
    for i, row in enumerate(rows[:35]):
        if any(c and "RPS" in str(c) and "-S" in str(c) for c in row):
            hdr_i = i
            break
    if hdr_i is None:
        return
    hdr = rows[hdr_i]
    models = {j: norm_id(c) for j, c in enumerate(hdr) if c and "RPS" in str(c)}
    start = hdr_i + 1
    for i in range(hdr_i + 1, min(hdr_i + 8, len(rows))):
        if rows[i][0] and str(rows[i][0]).lower() in ("feet", "tdh"):
            start = i + 1
            break
    store = {}
    for i in range(start, len(rows)):
        row = rows[i]
        if not row or row[0] is None:
            continue
        try:
            tdh = float(row[0])
        except (TypeError, ValueError):
            continue
        for j, mid in models.items():
            if j < len(row):
                try:
                    gpm = float(row[j])
                except (TypeError, ValueError):
                    continue
                add_point(store, mid, tdh, gpm)
    return store


def parse_model_grid(rows, header_pred=None, skip_s_suffix=True):
    """Find a model header row, then TDH rows below."""
    store = {}
    for i, row in enumerate(rows):
        if header_pred and not header_pred(row):
            continue
        models = {}
        for j, c in enumerate(row):
            if not c or "RPS" not in str(c):
                continue
            raw = str(c).strip()
            if skip_s_suffix and raw.replace(" ", "").endswith("-S"):
                continue
            mid = norm_id(c)
            if re.fullmatch(r"\d{2}RPS\d{2}", mid):
                models[j] = mid
        if len(models) < 3:
            continue
        for ri in range(i + 1, len(rows)):
            r = rows[ri]
            if not r or r[0] is None:
                continue
            try:
                tdh = float(r[0])
            except (TypeError, ValueError):
                if any(isinstance(r[j], str) and "MSRP" in str(r[j]) for j in range(len(r))):
                    break
                continue
            for j, mid in models.items():
                if j < len(r):
                    try:
                        gpm = float(r[j])
                    except (TypeError, ValueError):
                        continue
                    add_point(store, mid, tdh, gpm)
        if store:
            return store
    return store


def parse_st_xlsx(wb):
    store = {}
    for sheet in ST_SHEETS:
        rows = list(wb[sheet].iter_rows(values_only=True))
        merge = parse_model_grid(rows)
        for mid, pts in (merge or {}).items():
            store[mid] = dedupe_sort(pts)
    return store


def parse_xlsx(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    store = {}

    def merge_dict(d):
        if not d:
            return
        for mid, pts in d.items():
            store.setdefault(mid, []).extend(pts)

    for sheet in (
        "90150-S BIG RPS Sizing",
        "230300-S BIG RPS Sizing",
        "400500-S BIG RPS Sizing",
        "6508001100-S BIG RPS Sizing",
    ):
        rows = list(wb[sheet].iter_rows(values_only=True))
        merge_dict(parse_big_sheet(rows))

    rows6080 = list(wb["6080RPS Sizing"].iter_rows(values_only=True))
    merge_dict(
        parse_model_grid(
            rows6080,
            lambda r: any(c and str(c) in ("60RPS15", "80RPS50") for c in r),
        )
    )

    rows55 = list(wb["55RPS Sizing"].iter_rows(values_only=True))
    merge_dict(
        parse_model_grid(
            rows55, lambda r: any(c and str(c) == "55RPS10" for c in r)
        )
    )

    merge_dict(parse_st_xlsx(wb))
    wb.close()
    out = {mid: dedupe_sort(pts) for mid, pts in store.items()}
    return out


def merge_sources(md_data, xlsx_data):
    merged = dict(md_data)
    for mid, pts in xlsx_data.items():
        if mid in XLSX_WINS or mid.startswith(
            ("05RPS", "07RPS", "10RPS", "13RPS", "18RPS", "25RPS")
        ):
            if pts:
                merged[mid] = pts
        elif mid not in merged and pts:
            merged[mid] = pts
    return merged


def make_model(model_id, curve_data):
    head = model_id[-2:]
    color = "#cf8a7d" if model_id == "13RPS10" else HEAD_COLORS.get(head, "#888888")
    return {
        "id": model_id,
        "label": f"{model_id} ({HEAD_LABELS.get(head, head + ' HP')})",
        "color": color,
        "data": curve_data.get(model_id),
    }


def expand_st_families(doc, curve_data):
    fams = doc["families"]
    for key, heads in ST_FAMILY_HEADS.items():
        gpm = key[:2]
        ids = [f"{gpm}RPS{h}" for h in heads]
        fams[key]["models"] = [make_model(mid, curve_data) for mid in ids]

    ids_1318 = [f"{g}RPS{h}" for g, h in ST_13_18]
    fams["13_18RPS"]["models"] = [make_model(mid, curve_data) for mid in ids_1318]


def apply_to_families(families_doc, curve_data):
    updated = 0
    missing = []
    for key in families_doc["order"]:
        fam = families_doc["families"][key]
        for model in fam["models"]:
            mid = model["id"]
            if mid in curve_data and curve_data[mid]:
                model["data"] = curve_data[mid]
                updated += 1
            elif not model.get("data"):
                model["data"] = None
                missing.append(mid)
    return updated, missing


def main():
    md_data = parse_markdown(MD)
    xlsx_data = parse_xlsx(XLSX)
    merged = merge_sources(md_data, xlsx_data)

    doc = json.loads(FAM_JSON.read_text(encoding="utf-8"))
    expand_st_families(doc, merged)
    updated, still_null = apply_to_families(doc, merged)
    FAM_JSON.write_text(json.dumps(doc, indent=2) + "\n", encoding="utf-8")

    print(f"Markdown models: {len(md_data)}")
    print(f"XLSX models: {len(xlsx_data)}")
    print(f"Merged curves: {len(merged)}")
    print(f"Updated in families.json: {updated}")
    if still_null:
        print(f"Still line-only ({len(still_null)}): {', '.join(still_null)}")


if __name__ == "__main__":
    main()
